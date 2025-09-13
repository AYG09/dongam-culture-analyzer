"""
Excel (동암정신 7개요소.xlsx) -> dongam_spirit.json 병합 유틸리티
- 요구 사항:
  * 시트별(정신별) 표에서 행동/유형/무형 및 연결요소를 파싱
  * element_id 포맷은 '유형_#', '무형_#'로 정규화
  * 기존 JSON과 병합(업서트): 동일 spirit_id에 behaviors, tangible/intangible_elements, connected_elements 채움

사용 예시:
  python -m backend.modules.tools.xlsx_to_spirits --xlsx "public/동암정신 7개요소.xlsx" --json backend/modules/dongam_spirit.json --write

주의: 엑셀 포맷이 시트마다 조금씩 다를 수 있으므로, 열 헤더를 유연하게 탐지합니다.
"""
from __future__ import annotations
import argparse
import json
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
except Exception as e:  # pragma: no cover
    raise SystemExit("openpyxl가 필요합니다. backend/requirements.txt에 설치 후 다시 시도하세요.")


# ---------- Helpers ----------

def normalize_element_id(raw: str) -> str:
    if not raw:
        return raw
    raw = str(raw).strip()
    # 허용 입력: '유형1', '유형 1', '유형_1' 등 → '유형_1'
    #            '무형3', '무형-3', '무형_3' 등 → '무형_3'
    if raw.startswith("유형"):
        digits = ''.join(ch for ch in raw if ch.isdigit())
        return f"유형_{digits}" if digits else raw
    if raw.startswith("무형"):
        digits = ''.join(ch for ch in raw if ch.isdigit())
        return f"무형_{digits}" if digits else raw
    return raw


def split_ids(cell: Any) -> List[str]:
    if cell is None:
        return []
    s = str(cell).strip()
    if not s:
        return []
    # 쉼표/공백 구분 혼용 가정
    parts = [p for p in [x.strip() for x in s.replace(' ', ',').replace('\u3000', ',').split(',')] if p]
    return [normalize_element_id(p) for p in parts]


@dataclass
class Element:
    id: str
    name: str
    connected_elements: Optional[List[str]] = None

    def to_dict(self) -> Dict[str, Any]:
        d: Dict[str, Any] = {"id": self.id, "name": self.name}
        if self.connected_elements:
            d["connected_elements"] = self.connected_elements
        return d


@dataclass
class Behavior:
    id: str
    name: str
    connected_elements: Optional[List[str]] = None

    def to_dict(self) -> Dict[str, Any]:
        d: Dict[str, Any] = {"id": self.id, "name": self.name}
        if self.connected_elements:
            d["connected_elements"] = self.connected_elements
        return d


# ---------- Core Parsing ----------

def parse_sheet(ws) -> Tuple[List[Behavior], List[Element], List[Element]]:
    """시트 내 테이블에서 행동/유형/무형을 추출.
    엑셀 예시는 스크린샷 기준으로:
    - 첫 두 컬럼: [유형/행동 구분, 내용]
    - 오른쪽 8개 컬럼: 연결요소(행동1~8 헤더)
    실제 파일 헤더를 탐지해 유연히 처리한다.
    """
    # 헤더 탐지 (1~10행 탐색)
    header_row_idx = None
    headers: List[str] = []
    for r in range(1, 11):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
        joined = ''.join([str(x) for x in row_vals if x])
        if '연결요소' in joined or '행동' in joined:
            header_row_idx = r
            headers = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, 20)]
            break
    if header_row_idx is None:
        # 기본 가정: 1행이 헤더
        header_row_idx = 1
        headers = [str(ws.cell(row=1, column=c).value or '').strip() for c in range(1, 20)]

    # 연결요소 영역(열 인덱스) 추정: '연결요소' 또는 '행동'으로 시작하는 헤더 열들을 사용
    conn_cols: List[int] = []
    for idx, h in enumerate(headers, start=1):
        if h.startswith('연결요소') or h.startswith('행동'):
            conn_cols.append(idx)
    # 좌측 기본 컬럼 인덱스 추정
    kind_col = 1  # '유형/무형/행동' 구분
    content_col = 2  # 내용/이름

    behaviors: List[Behavior] = []
    tangible: List[Element] = []
    intangible: List[Element] = []

    # 본문 파싱
    ridx = header_row_idx + 1
    behavior_counter = 0
    id_counter_t = 0
    id_counter_i = 0
    while True:
        kind = ws.cell(row=ridx, column=kind_col).value
        name = ws.cell(row=ridx, column=content_col).value
        if kind is None and name is None:
            # 빈 행 연속 3개를 만나면 종료
            empty = 0
            for look_ahead in range(ridx, ridx + 3):
                if all(ws.cell(row=look_ahead, column=cc).value in (None, '') for cc in range(1, 6)):
                    empty += 1
            if empty >= 3:
                break
        kind_s = str(kind or '').strip()
        name_s = str(name or '').strip()

        # 연결요소 모으기
        conn: List[str] = []
        for cc in conn_cols:
            conn += split_ids(ws.cell(row=ridx, column=cc).value)
        conn = [normalize_element_id(x) for x in conn]
        conn = list(dict.fromkeys([x for x in conn if x]))  # uniq

        if not kind_s and not name_s:
            ridx += 1
            continue

        # 분류 규칙: '행동', '유형', '무형' 키워드로 판단
        if kind_s.startswith('행동') or kind_s.startswith('결과'):
            behavior_counter += 1
            behaviors.append(Behavior(id=f"행동{behavior_counter}", name=name_s or f"행동 {behavior_counter}", connected_elements=conn if conn else None))
        elif kind_s.startswith('유형'):
            id_counter_t += 1
            eid = normalize_element_id(kind_s)
            if not eid or eid == '유형_':
                eid = f'유형_{id_counter_t}'
            tangible.append(Element(id=eid, name=name_s, connected_elements=conn if conn else None))
        elif kind_s.startswith('무형'):
            id_counter_i += 1
            eid = normalize_element_id(kind_s)
            if not eid or eid == '무형_':
                eid = f'무형_{id_counter_i}'
            intangible.append(Element(id=eid, name=name_s, connected_elements=conn if conn else None))
        else:
            # 기타 행은 스킵
            pass

        ridx += 1

    return behaviors, tangible, intangible


def merge_into_json(base: Dict[str, Any], sheet_name: str, behaviors: List[Behavior], tangible: List[Element], intangible: List[Element]) -> Dict[str, Any]:
    # 시트명 → spirit_id 매핑 휴리스틱
    name_to_id = {
        '불우재': 'spirit_01',
        '숭조위선': 'spirit_02',
        '불굴의 도전정신과 개척정신': 'spirit_03',
        '미래를 예측하는 통찰': 'spirit_04',
        '미풍양속의 계승': 'spirit_05',
        '상생적 공존공영의 인화정신': 'spirit_06',
        '환경을 중시하는 사회적 책임경영': 'spirit_07',
    }
    target_id = None
    for k, v in name_to_id.items():
        if k in sheet_name:
            target_id = v
            break
    if target_id is None:
        # sheet_name이 id로 직접 되어 있으면 그대로 사용
        if sheet_name.startswith('spirit_'):
            target_id = sheet_name
        else:
            return base  # 매칭 실패 시 병합 건너뜀

    # 해당 spirit 찾기
    spirit = None
    for s in base.get('spirits', []):
        if s.get('id') == target_id:
            spirit = s
            break
    if spirit is None:
        return base

    # 병합 (덮어쓰기 기준 단순화: behaviors/tangible/intangible 요소를 통째로 갱신)
    if behaviors:
        spirit['behaviors'] = [b.to_dict() for b in behaviors]
    if tangible:
        spirit['tangible_elements'] = [e.to_dict() for e in tangible]
    if intangible:
        spirit['intangible_elements'] = [e.to_dict() for e in intangible]

    return base


# ---------- CLI ----------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True, help='엑셀 파일 경로 (예: public/동암정신 7개요소.xlsx)')
    ap.add_argument('--json', required=True, help='기존 JSON 파일 경로 (예: backend/modules/dongam_spirit.json)')
    ap.add_argument('--write', action='store_true', help='실제 파일에 덮어쓰기')
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"엑셀 파일을 찾을 수 없습니다: {args.xlsx}")
    if not os.path.exists(args.json):
        raise SystemExit(f"JSON 파일을 찾을 수 없습니다: {args.json}")

    wb = load_workbook(args.xlsx)
    with open(args.json, 'r', encoding='utf-8') as f:
        base = json.load(f)

    for ws in wb.worksheets:
        behaviors, tangible, intangible = parse_sheet(ws)
        base = merge_into_json(base, ws.title, behaviors, tangible, intangible)

    out_json = json.dumps(base, ensure_ascii=False, indent=2)
    if args.write:
        with open(args.json, 'w', encoding='utf-8') as f:
            f.write(out_json)
        print(f"UPDATED: {args.json}")
    else:
        print(out_json)


if __name__ == '__main__':
    main()
