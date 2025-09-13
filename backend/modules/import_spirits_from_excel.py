import json
import re
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
EXCEL_PATH = ROOT / "public" / "동암정신" / "동암정신 7개요소.xlsx"
JSON_PATH = Path(__file__).with_name("dongam_spirit.json")


def norm_id(value: str) -> str:
    """Normalize element_id to underscore format: '유형1' -> '유형_1', '무형14' -> '무형_14'.
    Accepts already normalized ids and returns them unchanged.
    """
    if not value:
        return value
    value = str(value).strip()
    # If already contains underscore, keep it
    if re.match(r"^(유형|무형)_[0-9]+$", value):
        return value
    m = re.match(r"^(유형|무형)\s*-?\s*([0-9]+)$", value)
    if m:
        return f"{m.group(1)}_{int(m.group(2))}"
    return value


def parse_sheet(ws) -> Dict[str, Any]:
    """Parse a sheet that follows the pattern shown in the screenshot:
    - Column A: 구분(결과/행동/유형/무형)
    - Column B: 내용(텍스트)
    - Columns starting around H..O: 연결요소 1..N with tokens like '행동1', '유형7', '무형13'

    Returns the structure with keys: behaviors, tangible_elements, intangible_elements.
    """
    behaviors: List[Dict[str, Any]] = []
    tangible: List[Dict[str, Any]] = []
    intangible: List[Dict[str, Any]] = []

    # Heuristics: locate header row by finding '연결요소' text
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and "연결요소" in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    start_row = 2 if header_row is None else header_row + 1

    # Find the first column index where header contains '연결요소'
    first_conn_col = None
    if header_row:
        for cell in ws[header_row]:
            if cell.value and "연결요소" in str(cell.value):
                first_conn_col = cell.column
                break

    # Fallback to a typical position if not found (col 8 = H)
    if first_conn_col is None:
        first_conn_col = 8

    # Scan rows
    for row in ws.iter_rows(min_row=start_row):
        kind = str(row[0].value).strip() if row[0].value else ""
        text = str(row[1].value).strip() if row[1].value else ""
        if not kind and not text:
            # stop if we hit a long empty area
            continue

        # Aggregate connections from the dynamic columns
        connections: List[str] = []
        for cell in row[first_conn_col - 1: first_conn_col + 8]:  # up to 8 connection columns
            val = cell.value
            if not val:
                continue
            sval = str(val).strip()
            # Accept patterns: '행동1', '유형7', '무형13', or '무형 13'
            m = re.match(r"^(행동|유형|무형)\s*-?\s*([0-9]+)$", sval)
            if m:
                token = f"{m.group(1)}_{int(m.group(2))}" if m.group(1) != "행동" else f"행동_{int(m.group(2))}"
                connections.append(token)

        if kind.startswith("행동"):
            # Determine number e.g., '행동5'
            m = re.match(r"^행동\s*-?\s*([0-9]+)$", kind)
            act_id = f"행동_{int(m.group(1))}" if m else f"행동_{len(behaviors)+1}"
            behaviors.append({
                "id": act_id,
                "name": text,
                "connected_elements": [norm_id(c) if c.startswith("유형") or c.startswith("무형") else c for c in connections],
            })
        elif kind.startswith("유형"):
            m = re.match(r"^유형\s*-?\s*([0-9]+)$", kind)
            el_id = f"유형_{int(m.group(1))}" if m else norm_id(kind)
            tangible.append({
                "id": el_id,
                "name": text,
                "connected_elements": [c for c in connections if c.startswith("무형_")],
            })
        elif kind.startswith("무형"):
            m = re.match(r"^무형\s*-?\s*([0-9]+)$", kind)
            el_id = f"무형_{int(m.group(1))}" if m else norm_id(kind)
            intangible.append({
                "id": el_id,
                "name": text,
                "connected_elements": [],
            })
        else:
            # Other kinds (결과 등) are ignored for card view
            continue

    # Deduplicate while preserving order
    def dedup(items: List[Dict[str, Any]], key: str) -> List[Dict[str, Any]]:
        seen = set()
        out = []
        for it in items:
            k = it.get(key)
            if k in seen:
                continue
            seen.add(k)
            out.append(it)
        return out

    return {
        "behaviors": dedup(behaviors, "id"),
        "tangible_elements": dedup(tangible, "id"),
        "intangible_elements": dedup(intangible, "id"),
    }


def merge_into_json(sheet_title: str, parsed: Dict[str, Any]) -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    # Find target spirit by simple title→id heuristic
    # Expect sheet title to contain the spirit's Korean name substring
    target = None
    for sp in data.get("spirits", []):
        if sp.get("name") and str(sheet_title).strip() in sp["name"]:
            target = sp
            break

    if not target:
        # Fallback match by contains (looser)
        for sp in data.get("spirits", []):
            if sp.get("name") and str(sp["name"]).find(str(sheet_title)) >= 0:
                target = sp
                break

    if not target:
        raise RuntimeError(f"No spirit matched for sheet '{sheet_title}'")

    # Merge fields
    target["behaviors"] = parsed.get("behaviors", [])
    target["tangible_elements"] = parsed.get("tangible_elements", [])
    target["intangible_elements"] = parsed.get("intangible_elements", [])

    # Save back
    JSON_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def import_all_sheets(xlsx_path: Path = EXCEL_PATH) -> List[str]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    updated: List[str] = []

    for sheet in wb.worksheets:
        title = sheet.title.strip()
        try:
            parsed = parse_sheet(sheet)
            # Only merge if some tangible/intangible are detected
            if parsed["tangible_elements"] or parsed["intangible_elements"]:
                merge_into_json(title, parsed)
                updated.append(title)
        except Exception as e:
            print(f"[WARN] Failed parsing sheet '{title}': {e}")

    return updated


def main():
    if not EXCEL_PATH.exists():
        print(f"Excel file not found: {EXCEL_PATH}")
        return
    updated = import_all_sheets(EXCEL_PATH)
    if updated:
        print("Updated sheets:", ", ".join(updated))
    else:
        print("No sheets updated. Check the Excel structure and headers (연결요소).")


if __name__ == "__main__":
    main()
